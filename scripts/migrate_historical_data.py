#!/usr/bin/env python3
import os
import sys
import datetime
import openpyxl
from pathlib import Path

# Add the parent directory of this script to the python path to load database.py
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.append(str(BASE_DIR))

try:
    from database import get_db_connection
except ImportError as e:
    print(f"Error importing get_db_connection: {e}")
    sys.exit(1)

MIGRATIONS_DIR = BASE_DIR / "migrations"

FILES_INFO = [
    {"name": "FEB-DATA-2026.xlsx", "month": 2, "year": 2026},
    {"name": "March-Data.xlsx", "month": 3, "year": 2026},
    {"name": "APRIL-DATA-2026.xlsx", "month": 4, "year": 2026},
    {"name": "JUNE-DATE-2026.xlsx", "month": 6, "year": 2026},
]

COLUMNS = [
    "compressor_suction_pressure",
    "compressor_discharge_pressure",
    "water_inlet_temp_c",
    "water_inlet_pressure_mpa",
    "water_outlet_temp_c",
    "water_outlet_pressure_mpa",
    "fan_amperes",
    "compressor_amperes",
    "ac_inlet_temp_in",
    "ac_inlet_temp_out",
    "ac_outlet_temp"
]

def create_table(conn):
    """Create the historical_data table in PostgreSQL if it doesn't exist."""
    print("Creating historical_data table if not exists...")
    with conn.cursor() as cur:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS historical_data (
                date DATE PRIMARY KEY,
                compressor_suction_pressure TEXT,
                compressor_discharge_pressure TEXT,
                water_inlet_temp_c REAL,
                water_inlet_pressure_mpa REAL,
                water_outlet_temp_c REAL,
                water_outlet_pressure_mpa REAL,
                fan_amperes REAL,
                compressor_amperes REAL,
                ac_inlet_temp_in REAL,
                ac_inlet_temp_out REAL,
                ac_outlet_temp REAL
            );
        """)
    conn.commit()
    print("Table created/verified successfully.")

def clean_value(val):
    """Clean the cell value, returning None if empty or NaN, otherwise string/float as appropriate."""
    if val is None:
        return None
    # If the value is a float/int, return it.
    if isinstance(val, (int, float)):
        # Check for NaN
        if val != val:  # NaN check
            return None
        return val
    # If string, strip it
    val_str = str(val).strip()
    if val_str.upper() in ["NAN", "NULL", "NONE", ""]:
        return None
    return val_str

def parse_excel_file(path, year, month):
    """Parse the Excel file and yield rows of data with dates."""
    print(f"Parsing {path.name}...")
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb.active

    # Row 1 and 2 are headers, actual data starts from Row 3
    for r in range(3, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 12)]
        
        # Clean the values
        cleaned_vals = [clean_value(v) for v in row_vals]

        # Skip completely empty rows
        if all(v is None for v in cleaned_vals):
            continue

        day = r - 2
        try:
            date_val = datetime.date(year, month, day)
        except ValueError:
            print(f"Warning: Day {day} is invalid for Month {month}, Year {year}. Skipping row {r}.")
            continue

        row_data = {"date": date_val}
        for col_name, val in zip(COLUMNS, cleaned_vals):
            row_data[col_name] = val
        
        yield row_data

def migrate_data():
    conn = get_db_connection()
    try:
        create_table(conn)
        
        total_inserted = 0
        total_updated = 0

        insert_sql = """
            INSERT INTO historical_data (
                date,
                compressor_suction_pressure,
                compressor_discharge_pressure,
                water_inlet_temp_c,
                water_inlet_pressure_mpa,
                water_outlet_temp_c,
                water_outlet_pressure_mpa,
                fan_amperes,
                compressor_amperes,
                ac_inlet_temp_in,
                ac_inlet_temp_out,
                ac_outlet_temp
            ) VALUES (
                %(date)s,
                %(compressor_suction_pressure)s,
                %(compressor_discharge_pressure)s,
                %(water_inlet_temp_c)s,
                %(water_inlet_pressure_mpa)s,
                %(water_outlet_temp_c)s,
                %(water_outlet_pressure_mpa)s,
                %(fan_amperes)s,
                %(compressor_amperes)s,
                %(ac_inlet_temp_in)s,
                %(ac_inlet_temp_out)s,
                %(ac_outlet_temp)s
            )
            ON CONFLICT (date) DO UPDATE SET
                compressor_suction_pressure = EXCLUDED.compressor_suction_pressure,
                compressor_discharge_pressure = EXCLUDED.compressor_discharge_pressure,
                water_inlet_temp_c = EXCLUDED.water_inlet_temp_c,
                water_inlet_pressure_mpa = EXCLUDED.water_inlet_pressure_mpa,
                water_outlet_temp_c = EXCLUDED.water_outlet_temp_c,
                water_outlet_pressure_mpa = EXCLUDED.water_outlet_pressure_mpa,
                fan_amperes = EXCLUDED.fan_amperes,
                compressor_amperes = EXCLUDED.compressor_amperes,
                ac_inlet_temp_in = EXCLUDED.ac_inlet_temp_in,
                ac_inlet_temp_out = EXCLUDED.ac_inlet_temp_out,
                ac_outlet_temp = EXCLUDED.ac_outlet_temp;
        """

        for info in FILES_INFO:
            path = MIGRATIONS_DIR / info["name"]
            if not path.exists():
                print(f"Warning: File {path.name} not found in migrations directory. Skipping.")
                continue

            rows_to_insert = list(parse_excel_file(path, info["year"], info["month"]))
            print(f"Parsed {len(rows_to_insert)} rows from {path.name}.")

            with conn.cursor() as cur:
                for row in rows_to_insert:
                    cur.execute(insert_sql, row)
            
            conn.commit()
            total_inserted += len(rows_to_insert)

        print(f"\nMigration completed successfully! Total records upserted: {total_inserted}")
    
    except Exception as e:
        conn.rollback()
        print(f"An error occurred during migration: {e}")
        sys.exit(1)
    finally:
        conn.close()

if __name__ == "__main__":
    migrate_data()
